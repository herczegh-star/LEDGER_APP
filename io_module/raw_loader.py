"""RAW Loader: načtení *_raw.xlsm / *_raw.csv → List[RawRow]. Žádná transformace významu."""
import csv
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Tuple
from core.model import RawRow


@dataclass
class LoadResult:
    """Výsledek načtení RAW souboru."""
    rows: List[RawRow] = field(default_factory=list)
    errors: List[dict] = field(default_factory=list)


def _parse_decimal(val) -> Optional[Decimal]:
    """Parsuje hodnotu na Decimal. Vrací None pro None, prázdný řetězec nebo nevalidní vstup."""
    if val is None or str(val).strip() == "":
        return None
    try:
        return Decimal(str(val).strip())
    except InvalidOperation:
        return None


def _parse_timestamp(val) -> datetime:
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ["%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%SZ"]:
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    raise ValueError(f"Nelze parsovat timestamp: {val}")


def _normalize_row(row_dict: dict, row_index: int = 0) -> Tuple[Optional[RawRow], Optional[dict]]:
    """Normalizuje dict na RawRow. Vrací (row, None) při úspěchu, (None, error_dict) při chybě."""
    errors = []

    try:
        ts = _parse_timestamp(row_dict.get("timestamp"))
    except ValueError as e:
        errors.append(str(e))
        ts = None

    amount = _parse_decimal(row_dict.get("amount"))
    if amount is None:
        errors.append(f"Nevalidní nebo chybějící amount: {row_dict.get('amount')!r}")
        amount = Decimal("0")

    price_raw = row_dict.get("price")
    price = None
    if price_raw is not None and str(price_raw).strip() != "":
        price = _parse_decimal(price_raw)
        if price is None:
            errors.append(f"Nevalidní price: {price_raw!r}")

    venue = str(row_dict.get("venue", "")).strip().lower()
    asset = str(row_dict.get("asset", "")).strip().upper()
    currency = str(row_dict.get("currency", "")).strip().upper()
    type_ = str(row_dict.get("type", "")).strip().upper()
    note = row_dict.get("note")
    if note is not None:
        note = str(note).strip()
        if note == "" or note.lower() == "nan" or note.lower() == "none":
            note = None
    row_id = row_dict.get("id")
    if row_id is not None:
        row_id = str(row_id).strip()
        if row_id == "" or row_id.lower() == "nan":
            row_id = None

    if errors:
        return None, {
            "row_index": row_index,
            "raw_data": {k: v for k, v in row_dict.items() if v is not None},
            "errors": errors,
        }

    return RawRow(
        id=row_id,
        timestamp=ts,
        type=type_,
        asset=asset,
        amount=amount,
        currency=currency,
        price=price,
        venue=venue,
        note=note,
    ), None


def load_xlsm(filepath: str) -> LoadResult:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "raw" in wb.sheetnames:
        ws = wb["raw"]
    else:
        ws = wb.active

    rows_data = list(ws.iter_rows(values_only=False))
    if not rows_data:
        return LoadResult()

    headers = [str(c.value).strip().lower() if c.value else "" for c in rows_data[0]]
    result = LoadResult()
    for i, row in enumerate(rows_data[1:], start=1):
        vals = [c.value for c in row]
        row_dict = dict(zip(headers, vals))
        if not row_dict.get("type") or not row_dict.get("asset"):
            continue
        if str(row_dict.get("type", "")).strip() == "":
            continue
        parsed, error = _normalize_row(row_dict, row_index=i)
        if error:
            result.errors.append(error)
        else:
            result.rows.append(parsed)
    return result


def load_csv(filepath: str) -> LoadResult:
    result = LoadResult()
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i, row_dict in enumerate(reader, start=1):
            if not row_dict.get("type") or not row_dict.get("asset"):
                continue
            parsed, error = _normalize_row(row_dict, row_index=i)
            if error:
                result.errors.append(error)
            else:
                result.rows.append(parsed)
    return result


def load_raw(filepath: str) -> LoadResult:
    path = Path(filepath)
    if path.suffix in (".xlsm", ".xlsx"):
        return load_xlsm(filepath)
    elif path.suffix == ".csv":
        return load_csv(filepath)
    else:
        raise ValueError(f"Nepodporovaný formát: {path.suffix}")

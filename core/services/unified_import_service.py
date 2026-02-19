"""Unified format raw importer.

Reads .csv / .xlsx / .xlsm files already in unified_format_raw schema and
appends rows to the ledger via LedgerStore.import_rows() (dedup-safe).

Does NOT map / transform vendor-specific formats.  The input file must already
match the canonical schema.  For vendor files use io_module.raw_loader instead.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import FrozenSet, List, Optional

from core.ledger_store import LedgerStore
from core.model import RawRow
from io_module.raw_loader import _normalize_row  # shared parse + normalise logic

# ── Column requirements ────────────────────────────────────────────────────────

_REQUIRED_COLS: FrozenSet[str] = frozenset(
    {"timestamp", "type", "asset", "amount", "currency", "venue"}
)


def _check_columns(headers: set, path: str) -> None:
    """Raise ValueError if any required column is absent from *headers*."""
    missing = _REQUIRED_COLS - headers
    if missing:
        raise ValueError(
            f"{Path(path).name!r} is missing required columns: {sorted(missing)}. "
            f"Required: {sorted(_REQUIRED_COLS)}"
        )


# ── File readers → list[dict] ─────────────────────────────────────────────────

def _read_csv(path: str) -> List[dict]:
    """Read CSV into a list of lowercase-keyed dicts.  Validates required columns."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        raw_headers = {h.strip().lower() for h in (reader.fieldnames or [])}
        _check_columns(raw_headers, path)
        return [
            {k.strip().lower(): v for k, v in row.items()}
            for row in reader
        ]


def _read_xlsx(path: str, sheet_name: Optional[str]) -> List[dict]:
    """Read XLSX/XLSM sheet into a list of lowercase-keyed dicts.

    Sheet priority:
      1. *sheet_name* if explicitly provided
      2. Sheet named "raw" (matches existing raw_loader convention)
      3. Active (first) sheet
    """
    import openpyxl  # optional dependency; imported lazily

    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {Path(path).name!r}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    elif "raw" in wb.sheetnames:
        ws = wb["raw"]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        return []

    headers = [
        str(h).strip().lower() if h is not None else ""
        for h in raw_headers
    ]
    _check_columns({h for h in headers if h}, path)

    return [
        {headers[i]: cell for i, cell in enumerate(row)}
        for row in rows_iter
    ]


# ── Public API ────────────────────────────────────────────────────────────────

def import_unified_file(
    db_path: str,
    file_path: str,
    sheet_name: Optional[str] = None,
) -> List[RawRow]:
    """Load unified_format_raw rows from a file and append to the ledger.

    Args:
        db_path:    Path to the SQLite ledger DB.
        file_path:  Path to a .csv, .xlsx, or .xlsm file.
        sheet_name: XLSX sheet name override (default: "raw" sheet, then active sheet).

    Returns:
        List of RawRow objects parsed from the file.
        Rows that are exact duplicates of existing ledger entries are silently
        skipped by LedgerStore (fingerprint deduplication) but still returned.

    Raises:
        ValueError: Missing required columns, unsupported file format, or parse
                    errors in one or more data rows.
    """
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        raw_dicts = _read_csv(file_path)
    elif ext in (".xlsx", ".xlsm"):
        raw_dicts = _read_xlsx(file_path, sheet_name)
    else:
        raise ValueError(
            f"Unsupported file format {ext!r}. Supported: .csv, .xlsx, .xlsm"
        )

    rows: List[RawRow] = []
    parse_errors: List[str] = []

    for i, row_dict in enumerate(raw_dicts, start=2):  # row 1 = header
        # Skip entirely blank rows (e.g. trailing empty lines in XLSX)
        type_val = str(row_dict.get("type") or "").strip()
        asset_val = str(row_dict.get("asset") or "").strip()
        if not type_val and not asset_val:
            continue

        parsed, error = _normalize_row(row_dict, row_index=i)
        if error:
            parse_errors.append(f"Row {i}: {'; '.join(error['errors'])}")
        else:
            rows.append(parsed)

    if parse_errors:
        raise ValueError(
            f"{len(parse_errors)} row(s) failed to parse:\n" + "\n".join(parse_errors)
        )

    if not rows:
        return []

    store = LedgerStore(db_path)
    try:
        store.import_rows(rows)
    finally:
        store.close()

    return rows
